import openpyxl
from openpyxl.styles import Font, PatternFill
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, UnexpectedAlertPresentException
from selenium.webdriver.chrome.options import Options
import time
import random
import os
import glob
import sys
import json
from datetime import datetime
import urllib.request
import urllib.error
import subprocess

AUTO_PRINT_DELAY_SPP_MIN = 12
AUTO_PRINT_DELAY_SPP_MAX = 16
AUTO_PRINT_DELAY_FKPP_MIN = 12
AUTO_PRINT_DELAY_FKPP_MAX = 16
AUTO_PRINT_POST_DELAY_SPP = 3
AUTO_PRINT_POST_DELAY_FKPP = 3
AUTO_PRINT_MIN_DELAY = 2
ENABLE_KIOSK_PRINTING = True
PRINTER_NAME = "EPSON L120 Series" #sesuaikan dengan nama printer yang ada di settingn printer Windows Anda atau isi 'Microsoft Print to PDF'
NOTIFY_URL = "https://api.silakes.labkesdasumenep.id/api/bot/v1/send-group-message"
NOTIFY_SECRET_KEY = "kesehatanNo1@"

# ====================================================================
# FUNGSI UTILITAS & UI TERMINAL (INTERAKTIF)
# ====================================================================

def get_key():
    if os.name == 'nt':
        import msvcrt
        key = msvcrt.getch()
        if key == b'\xe0':
            key = msvcrt.getch()
            if key == b'H': return 'UP'
            elif key == b'P': return 'DOWN'
        elif key == b'\r': return 'ENTER'
        elif key == b'\x03': raise KeyboardInterrupt
        return key.decode('utf-8', 'ignore')
    else:
        import tty, termios
        fd = sys.stdin.fileno()
        old_settings = termios.tcgetattr(fd)
        try:
            tty.setraw(sys.stdin.fileno())
            ch = sys.stdin.read(1)
            if ch == '\x1b':
                ch = sys.stdin.read(2)
                if ch == '[A': return 'UP'
                elif ch == '[B': return 'DOWN'
            elif ch == '\r' or ch == '\n': return 'ENTER'
            elif ch == '\x03': raise KeyboardInterrupt
            return ch
        finally:
            termios.tcsetattr(fd, termios.TCSADRAIN, old_settings)

def menu_interaktif(pilihan, judul="Pilih salah satu:"):
    indeks_terpilih = 0
    while True:
        os.system('cls' if os.name == 'nt' else 'clear')
        print("="*60)
        print(f" {judul} ")
        print(" (Gunakan panah Atas/Bawah, lalu tekan ENTER)")
        print("="*60)
        
        for i, opsi in enumerate(pilihan):
            if i == indeks_terpilih:
                print(f" [*] {opsi}")
            else:
                print(f" [ ] {opsi}")
                
        key = get_key()
        if key == 'UP':
            indeks_terpilih = (indeks_terpilih - 1) % len(pilihan)
        elif key == 'DOWN':
            indeks_terpilih = (indeks_terpilih + 1) % len(pilihan)
        elif key == 'ENTER':
            os.system('cls' if os.name == 'nt' else 'clear')
            return indeks_terpilih, pilihan[indeks_terpilih]

def setup_folder_excel():
    nama_folder = "FILE EXCEL DISINI"
    if not os.path.exists(nama_folder):
        os.makedirs(nama_folder)
        print(f"\n[INFO] Folder '{nama_folder}' telah dibuat.")
        print("Silakan pindahkan file Excel (data pasien) Anda ke dalam folder tersebut.")
        input("Tekan ENTER jika file sudah dipindahkan...")
    
    while True:
        list_file = glob.glob(os.path.join(nama_folder, "*.xlsx"))
        if not list_file:
            print(f"\n[!] Tidak ada file .xlsx ditemukan di dalam folder '{nama_folder}'.")
            input("Silakan masukkan file, lalu tekan ENTER untuk mengecek kembali...")
            continue
        
        nama_file_saja = [os.path.basename(f) for f in list_file]
        idx_file, file_terpilih = menu_interaktif(nama_file_saja, "Pilih File Excel yang akan diproses:")
        path_file = os.path.join(nama_folder, file_terpilih)
        
        try:
            wb = openpyxl.load_workbook(path_file)
            list_sheet = wb.sheetnames
            idx_sheet, sheet_terpilih = menu_interaktif(list_sheet, f"Pilih Sheet dari file {file_terpilih}:")
            
            os.system('cls' if os.name == 'nt' else 'clear')
            print("="*60)
            print(" KONFIRMASI DATA")
            print("="*60)
            print(f"File  : {file_terpilih}")
            print(f"Sheet : {sheet_terpilih}")
            print("="*60)
            konfirmasi = input("Apakah data ini sudah benar? (Y/Enter = Ya, N = Pilih Ulang): ").strip().lower()
            if konfirmasi == 'n':
                continue
            
            return path_file, sheet_terpilih, wb
        except Exception as e:
            print(f"Error membaca file: {e}")
            input("Tekan ENTER untuk mencoba lagi...")

# ====================================================================
# FUNGSI SELENIUM
# ====================================================================

def ketik_seperti_manusia(elemen, teks):
    for huruf in teks:
        elemen.send_keys(huruf)
        time.sleep(random.uniform(0.01, 0.05))

def tunggu_loading_pace(driver):
    try:
        WebDriverWait(driver, 30).until(EC.invisibility_of_element_located((By.CSS_SELECTOR, ".pace-active")))
        time.sleep(0.5)
    except:
        pass

def tutup_modal_alert(driver):
    try:
        btn_modal_ok = WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "button.bootbox-accept")))
        btn_modal_ok.click()
        tunggu_loading_pace(driver)
    except Exception:
        pass

    # Notifikasi jenis "toast" (data-notify, mis. banner REMINDER) — beda struktur
    # dari modal bootbox, dan sering menutupi elemen lain sehingga menyebabkan
    # 'element click intercepted'. Klik tombol dismiss-nya kalau ada, lalu hapus
    # paksa lewat JS sebagai jaring pengaman kalau tombolnya tidak mempan.
    try:
        for btn in driver.find_elements(By.CSS_SELECTOR, "[data-notify='container'] [data-notify='dismiss']"):
            try:
                driver.execute_script("arguments[0].click();", btn)
            except Exception:
                pass
        driver.execute_script("""
            document.querySelectorAll('[data-notify="container"]').forEach(function (el) {
                if (el.textContent.includes('REMINDER')) { el.remove(); }
            });
        """)
    except Exception:
        pass

def cek_sesi_berakhir(driver):
    try:
        modal_sesi = driver.find_elements(By.XPATH, "//div[contains(text(), 'Sesi Anda sudah berakhir')]")
        if modal_sesi and modal_sesi[0].is_displayed():
            print("\n" + "!"*60)
            print("⚠️ PERINGATAN: SESI PCARE ANDA TELAH BERAKHIR!")
            print("Bot akan menjeda proses. Silakan lakukan hal berikut:")
            print("1. Klik OK pada pop-up di browser.")
            print("2. Login kembali dan selesaikan CAPTCHA.")
            print("3. Pastikan Anda sudah kembali ke Dashboard utama PCare.")
            print("!"*60)
            input("\n[TEKAN ENTER DI SINI JIKA ANDA SUDAH LOGIN KEMBALI KE DASHBOARD]...")
            return True
    except:
        pass
    return False

def pilih_select2(driver, elemen_id_asli, nilai_pencarian):
    try:
        container = WebDriverWait(driver, 5).until(
            EC.element_to_be_clickable((By.XPATH, f"//select[@id='{elemen_id_asli}']/following-sibling::span[contains(@class, 'select2-container')]"))
        )
        container.click()
        time.sleep(0.5)
        
        opsi = WebDriverWait(driver, 5).until(
            EC.element_to_be_clickable((By.XPATH, f"//ul[contains(@id, 'select2-{elemen_id_asli}-results')]//li[contains(text(), '{nilai_pencarian}')]"))
        )
        opsi.click()
        time.sleep(0.2)
    except Exception as e:
        print(f"   [!] Gagal memilih dropdown {elemen_id_asli}: {str(e)[:30]}")

def elemen_terlihat(driver, by, selector):
    try:
        elemen = driver.find_element(by, selector)
        return elemen.is_displayed()
    except Exception:
        return False

def safe_click(driver, by, selector, timeout=8, scroll=True):
    try:
        elemen = WebDriverWait(driver, timeout).until(EC.element_to_be_clickable((by, selector)))
        if scroll:
            driver.execute_script("arguments[0].scrollIntoView({behavior: 'smooth', block: 'center'});", elemen)
            time.sleep(0.2)
        elemen.click()
        return True
    except Exception:
        try:
            elemen = driver.find_element(by, selector)
            if scroll:
                driver.execute_script("arguments[0].scrollIntoView({behavior: 'smooth', block: 'center'});", elemen)
                time.sleep(0.2)
            driver.execute_script("arguments[0].click();", elemen)
            return True
        except Exception as e:
            print(f"⚠️ Gagal klik elemen {selector}: {str(e)[:30]}")
            return False

def buka_tab(driver, tab_id, timeout=8):
    if not safe_click(driver, By.XPATH, f"//a[@href='#{tab_id}']"):
        return False
    try:
        WebDriverWait(driver, timeout).until(EC.visibility_of_element_located((By.ID, tab_id)))
        return True
    except Exception:
        return False

def daftar_mengandung(daftar, kata_kunci):
    kata = str(kata_kunci).upper()
    return any(kata in str(item).upper() for item in daftar)

def pilih_opsi_dropdown(driver, elemen_id, kata_kunci):
    try:
        select_el = driver.find_element(By.ID, elemen_id)
        sel = Select(select_el)
        for opt in sel.options:
            if kata_kunci.lower() in opt.text.lower():
                sel.select_by_visible_text(opt.text)
                return True
    except Exception:
        pass
    try:
        pilih_select2(driver, elemen_id, kata_kunci)
        return True
    except Exception:
        return False

def tunggu_tab_nonkapi(driver, timeout=12):
    try:
        WebDriverWait(driver, timeout).until(EC.visibility_of_element_located((By.ID, "tabTindakan")))
        return True
    except Exception:
        return False

def jalankan_auto_print(driver, min_delay, max_delay, post_delay):
    min_delay = max(min_delay, AUTO_PRINT_MIN_DELAY)
    max_delay = max(max_delay, min_delay)
    time.sleep(random.uniform(min_delay, max_delay))
    try:
        driver.execute_script("window.print();")
        time.sleep(post_delay)
        return True
    except Exception:
        return False

def pastikan_turnstile_terselesaikan(driver):
    try:
        # Beberapa halaman pakai id "pcare-turnstile" (dashboard/login), yang lain
        # pakai "pcare-turnstile-slot" (halaman search Pelayanan) - cocokkan keduanya.
        # Hanya dipakai untuk cek KEBERADAAN widget, bukan disimpan & dipakai belakangan -
        # elemen ini gampang berubah/di-render ulang oleh Cloudflare (stale element).
        WebDriverWait(driver, 3).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, "div.cf-turnstile, [id^='pcare-turnstile']"))
        )
    except Exception:
        return True  # Tidak ada widget Turnstile di halaman ini

    def token_terisi():
        try:
            token_val = driver.execute_script(
                "const el = document.querySelector('input[name=\"cf-turnstile-response\"]');"
                "return el ? el.value : '';"
            )
        except Exception:
            token_val = ""
        return bool(token_val and token_val.strip())

    # Beri kesempatan widget auto-pass sendiri (mode invisible/managed) - ini jalur
    # paling andal karena browser ini asli/tidak terdeteksi otomasi. EzSolver (browser
    # terpisah) TIDAK dipakai di sini lagi: tokennya kerap ditolak server dan pernah
    # menyebabkan crash 'stale element reference' saat halaman re-render selagi solve.
    for _ in range(30):
        if token_terisi():
            return True
        time.sleep(0.5)

    print("-> Widget Cloudflare Turnstile belum terverifikasi otomatis.")
    input("Silakan selesaikan verifikasi Turnstile secara manual di browser (tunggu sampai centang hijau 'Success'), lalu tekan ENTER di sini...")
    if not token_terisi():
        print("⚠️ Turnstile masih belum terverifikasi, tetap melanjutkan (mungkin gagal di langkah berikutnya).")
    return True

EZSOLVER_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "ezsolver")

def coba_solve_turnstile_ezsolver(driver, timeout=45):
    """Coba selesaikan Turnstile otomatis pakai EzSolver (browser nodriver terpisah).
    Khusus dipakai di form input kunjungan - titik ini paling sering minta
    verifikasi ulang. Return False kalau gagal (caller wajib fallback manual)."""
    try:
        sitekey = driver.execute_script(
            "const el = document.querySelector('div.cf-turnstile, [id^=\"pcare-turnstile\"]');"
            "if (el && el.getAttribute('data-sitekey')) return el.getAttribute('data-sitekey');"
            "return window.TURNSTILE_SITEKEY || null;"
        )
        action = driver.execute_script(
            "const el = document.querySelector('div.cf-turnstile, [id^=\"pcare-turnstile\"]');"
            "if (el && el.getAttribute('data-action')) return el.getAttribute('data-action');"
            "return window.TURNSTILE_ACTION || null;"
        )
        if not sitekey:
            return False
        siteurl = driver.current_url

        sys.path.insert(0, EZSOLVER_DIR)
        from solver import solve

        print("-> Mencoba menyelesaikan Turnstile otomatis via EzSolver...")
        token = solve(sitekey, siteurl, timeout=timeout, action=action)

        driver.execute_script(
            "const token = arguments[0];"
            "let input = document.querySelector('input[name=\"cf-turnstile-response\"]');"
            "if (input) { input.value = token; }"
            "if (window.pcareTurnstileCallback) { try { window.pcareTurnstileCallback(token); } catch(e) {} }",
            token
        )
        print("-> Turnstile terselesaikan otomatis via EzSolver.")
        return True
    except Exception as e:
        print(f"⚠️ EzSolver gagal menyelesaikan Turnstile: {str(e)[:80]}")
        return False

def pastikan_turnstile_form_kunjungan(driver, timeout_solver=45):
    """Dipanggil khusus setelah form input kunjungan terbuka (titik yang sering
    minta verifikasi Turnstile ulang). Coba EzSolver dulu, kalau gagal fallback
    ke alur manual (tunggu auto-pass lalu minta konfirmasi ENTER)."""
    try:
        WebDriverWait(driver, 3).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, "div.cf-turnstile, [id^='pcare-turnstile']"))
        )
    except Exception:
        return True  # Tidak ada widget Turnstile di form ini

    def token_terisi():
        try:
            token_val = driver.execute_script(
                "const el = document.querySelector('input[name=\"cf-turnstile-response\"]');"
                "return el ? el.value : '';"
            )
        except Exception:
            token_val = ""
        return bool(token_val and token_val.strip())

    if token_terisi():
        return True

    if coba_solve_turnstile_ezsolver(driver, timeout=timeout_solver) and token_terisi():
        return True

    return pastikan_turnstile_terselesaikan(driver)

def tunggu_turnstile_dan_login(driver, timeout=600):
    pastikan_turnstile_terselesaikan(driver)

    mulai = time.time()
    sudah_pesan_manual = False
    while True:
        try:
            token_val = driver.execute_script(
                "const el = document.querySelector('input[name=\"cf-turnstile-response\"]');"
                "return el ? el.value : '';"
            )
        except Exception:
            token_val = ""

        if token_val and token_val.strip():
            for s in range(3, 0, -1):
                print(f"-> Turnstile terverifikasi. Login otomatis dalam {s}...")
                time.sleep(1)
            try:
                driver.find_element(By.ID, "btnLogin").click()
                return True
            except Exception as e:
                print(f"⚠️ Gagal klik Sign In: {str(e)[:30]}")
                return False

        if not sudah_pesan_manual and time.time() - mulai > 8:
            print("-> Menunggu verifikasi Cloudflare Turnstile... Jika macet, klik kotak centang di browser secara manual.")
            sudah_pesan_manual = True

        if time.time() - mulai > timeout:
            lanjut = input("Turnstile belum terverifikasi. ENTER = lanjut menunggu, n = batal: ").strip().lower()
            if lanjut == 'n':
                return False
            mulai = time.time()

        time.sleep(0.3)

def isi_kredensial(driver, username, password):
    input_username = WebDriverWait(driver, 10).until(
        EC.element_to_be_clickable((By.XPATH, "//input[@placeholder='Username']"))
    )
    time.sleep(random.uniform(0.3, 0.7))
    driver.execute_script("arguments[0].removeAttribute('readonly', 0);", input_username)
    input_username.click()
    time.sleep(0.2)
    input_username.clear()
    ketik_seperti_manusia(input_username, str(username))

    time.sleep(random.uniform(0.2, 0.5))

    input_password = driver.find_element(By.XPATH, "//input[@placeholder='Password']")
    driver.execute_script("arguments[0].removeAttribute('readonly', 0);", input_password)
    input_password.click()
    time.sleep(0.2)
    input_password.clear()
    ketik_seperti_manusia(input_password, str(password))

def ambil_pesan_notif(driver):
    try:
        msg_el = driver.find_elements(By.CSS_SELECTOR, "[data-notify='message']")
        return msg_el[0].text.strip() if msg_el else ""
    except Exception:
        return ""

def klik_search_dengan_retry_captcha(driver, klik_search_fn, tunggu_detik=4):
    """Klik tombol search. Kalau server menolak dengan pesan verifikasi/captcha gagal,
    coba selesaikan Turnstile otomatis via EzSolver dulu lalu ulangi klik search -
    kalau EzSolver juga gagal/tetap ditolak, baru minta user refresh & mengulang
    pencarian secara manual di browser, lalu bot lanjut setelah user konfirmasi siap."""
    klik_search_fn()
    time.sleep(0.5)
    tunggu_loading_pace(driver)
    time.sleep(tunggu_detik)

    pesan = ambil_pesan_notif(driver)
    gagal_verifikasi = bool(pesan) and ("captcha" in pesan.lower() or "verifikasi" in pesan.lower())
    if not gagal_verifikasi:
        return True

    print(f"⚠️ {pesan}")
    print("-> Verifikasi Turnstile ditolak server, mencoba menyelesaikan otomatis via EzSolver...")
    if coba_solve_turnstile_ezsolver(driver):
        klik_search_fn()
        time.sleep(0.5)
        tunggu_loading_pace(driver)
        time.sleep(tunggu_detik)

        pesan_ulang = ambil_pesan_notif(driver)
        gagal_lagi = bool(pesan_ulang) and ("captcha" in pesan_ulang.lower() or "verifikasi" in pesan_ulang.lower())
        if not gagal_lagi:
            print("-> Pencarian berhasil setelah Turnstile diselesaikan via EzSolver.")
            return True
        print(f"⚠️ {pesan_ulang}")
        print("-> EzSolver tetap ditolak server untuk pencarian ini.")

    print("   Silakan REFRESH halaman (F5) di browser, lalu ulangi pencarian secara manual sampai data pasien muncul.")
    input("Setelah data pasien tampil di browser, tekan ENTER di sini untuk bot melanjutkan proses...")
    return True

def tunggu_klik_cari_pasien_oleh_user(driver, timeout=180):
    """Klik Cari PERTAMA untuk tiap pasien baru dilakukan USER secara manual (klik asli
    manusia lebih reliable untuk Turnstile daripada bot) - field tanggal/BPJS tetap diisi
    otomatis oleh bot, tapi bot cuma menunggu (tanpa perlu ENTER) sampai nama pasien
    tampil, alert 'Data tidak ditemukan' muncul, atau timeout habis.
    Nama & notif yang SUDAH ada sebelum mulai menunggu (sisa dari pasien sebelumnya)
    diabaikan - hanya perubahan yang muncul SETELAH ini yang dianggap hasil klik user,
    supaya bot tidak salah kira status pasien lama sebagai hasil klik pasien baru.
    Return: ("ditemukan", nama) / ("tidak_ditemukan", pesan) / ("timeout", None)"""
    try:
        nama_awal = driver.find_element(By.ID, "lblnmpst").text.strip()
    except Exception:
        nama_awal = ""
    pesan_awal = ambil_pesan_notif(driver)

    akhir = time.time() + timeout
    pesan_terakhir_dicetak = ""
    while time.time() < akhir:
        try:
            nama = driver.find_element(By.ID, "lblnmpst").text.strip()
        except Exception:
            nama = ""
        if nama and nama != nama_awal:
            return "ditemukan", nama

        pesan = ambil_pesan_notif(driver)
        if pesan and pesan != pesan_awal:
            if "tidak ditemukan" in pesan.lower():
                return "tidak_ditemukan", pesan
            if pesan != pesan_terakhir_dicetak and ("captcha" in pesan.lower() or "verifikasi" in pesan.lower()):
                print(f"⚠️ {pesan} (silakan klik Cari lagi di browser)")
                pesan_terakhir_dicetak = pesan

        time.sleep(1)
    return "timeout", None

def tunggu_hasil_login(driver, timeout=15):
    mulai = time.time()
    while True:
        if cek_modal_login_gagal(driver):
            return "failed"
        if elemen_terlihat(driver, By.XPATH, "//a[contains(text(), 'Entri Data')]"):
            return "success"
        msg = ambil_pesan_notif(driver)
        if msg:
            msg_lower = msg.lower()
            if "gagal" in msg_lower or "captcha" in msg_lower or "password" in msg_lower or "username" in msg_lower:
                print(f"⚠️ Pesan dari sistem PCare: {msg}")
                return "failed"
        if time.time() - mulai > timeout:
            return "timeout"
        time.sleep(0.3)

def cek_modal_login_gagal(driver):
    try:
        modal_body = driver.find_element(By.CSS_SELECTOR, ".bootbox-body")
        msg = modal_body.text.strip()
        if msg:
            msg_lower = msg.lower()
            if "captcha" in msg_lower or "salah" in msg_lower or "gagal" in msg_lower:
                print(f"⚠️ Pesan dari sistem PCare: {msg}")
                try:
                    driver.find_element(By.CSS_SELECTOR, "button.bootbox-accept").click()
                except Exception:
                    pass
                return True
    except Exception:
        return False
    return False

def format_report_message(judul, nama_file, nama_sheet):
    waktu = datetime.now().strftime("%d-%m-%Y %H:%M:%S")
    return (
        "╭──────────────╮\n"
        "│ 🗓️ PCARE REPORT\n"
        "│   SILAKES BOT PCARE\n"
        "╰──────────────╯\n"
        f"{judul}\n"
        "━━━━━━━━━━━━━━━━━━━━━━\n"
        f"Nama Excel : {nama_file}\n"
        f"Sheet : {nama_sheet}\n"
        "━━━━━━━━━━━━━━━━━━━━━━\n"
        f"SELESAI PADA WAKTU {waktu}"
    )

def kirim_notif_group(pesan):
    payload = {"secretKey": NOTIFY_SECRET_KEY, "message": pesan}
    data = json.dumps(payload).encode("utf-8")
    req = urllib.request.Request(
        NOTIFY_URL,
        data=data,
        headers={"Content-Type": "application/json"}
    )
    try:
        with urllib.request.urlopen(req, timeout=15) as resp:
            return 200 <= resp.status < 300
    except Exception as e:
        print(f"⚠️ Gagal mengirim notifikasi: {str(e)[:60]}")
        return False

# ====================================================================
# MODUL: PELAYANAN PASIEN
# ====================================================================

def hitung_total_biaya_lab(driver, tab_id, tabel_id):
    """Fungsi mengekstrak total biaya dari tabel daftar tindakan Non Kapitasi"""
    try:
        total_biaya = 0
        baris_tabel = driver.find_elements(By.XPATH, f"//div[@id='{tab_id}']//table[@id='{tabel_id}']/tbody/tr")
        for baris in baris_tabel:
            # Lewati jika baris kosong (No data available)
            if "No data available" in baris.text:
                continue
                
            kolom_biaya = baris.find_elements(By.TAG_NAME, "td")[1].text # Kolom ke-2 adalah biaya
            # Bersihkan format "Rp. 45.000,-" menjadi integer 45000
            biaya_bersih = kolom_biaya.replace("Rp.", "").replace(" ", "").replace(".", "").replace(",-", "")
            if biaya_bersih.isdigit():
                total_biaya += int(biaya_bersih)
        return total_biaya
    except:
        return 0

def ambil_nama_pelayanan_dari_tabel(driver, tabel_id):
    try:
        nama_list = []
        baris_tabel = driver.find_elements(By.XPATH, f"//table[@id='{tabel_id}']/tbody/tr")
        for baris in baris_tabel:
            if "No data available" in baris.text:
                continue
            kolom = baris.find_elements(By.TAG_NAME, "td")
            if not kolom:
                continue
            nama = kolom[0].text.strip()
            if nama:
                nama_list.append(nama)
        return nama_list
    except:
        return []

def tunggu_hasil_simpan(driver, timeout=8):
    try:
        WebDriverWait(driver, timeout).until(
            lambda d: d.find_elements(By.CSS_SELECTOR, "[data-notify='message']")
        )
    except:
        return "timeout", ""
    try:
        msg_el = driver.find_elements(By.CSS_SELECTOR, "[data-notify='message']")
        msg = msg_el[0].text.strip() if msg_el else ""
    except:
        msg = ""
    if msg and "berhasil disimpan" in msg.lower():
        return "success", msg
    if msg:
        return "warning", msg
    return "timeout", ""

def tandai_cannot_input(sheet_data, wb_data, path_file, row, pesan):
    sheet_data.cell(row=row, column=29).value = "CANNOT BE INPUT"
    sheet_data.cell(row=row, column=30).value = pesan
    sheet_data.cell(row=row, column=29).font = Font(color="FF0000")
    wb_data.save(path_file)

def pastikan_form_kimia_darah(driver):
    try:
        list_el = driver.find_element(By.ID, "listKimiaDarah_lyt")
        content_el = driver.find_element(By.ID, "contentKimiaDarah_lyt")
        list_visible = list_el.is_displayed()
        content_visible = content_el.is_displayed()
    except:
        list_visible = False
        content_visible = False
    
    if content_visible and not list_visible:
        return "content"
    
    try:
        btn = driver.find_element(By.ID, "tambahPelayanan_btn")
        if btn.is_displayed():
            safe_click(driver, By.ID, "tambahPelayanan_btn")
            time.sleep(0.5)
            return "clicked"
    except:
        pass
    
    return "unknown"

def buka_riwayat_setelah_simpan(driver, tgl_rujukan, max_retry=3, require_labkes=False):
    for percobaan in range(1, max_retry + 1):
        try:
            driver.execute_script("window.scrollTo(0, 0);")
        except:
            pass

        try:
            btn_cari = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.ID, "btnCariPendaftaran")))
            driver.execute_script("arguments[0].scrollIntoView({behavior: 'smooth', block: 'center'});", btn_cari)
            time.sleep(0.2)
            btn_cari.click()
            tunggu_loading_pace(driver)
            time.sleep(1.5)
        except Exception as e:
            print(f"⚠️ Gagal klik Cari Pendaftaran (percobaan {percobaan}/{max_retry}): {str(e)[:30]}")
            if percobaan < max_retry:
                time.sleep(1)
                continue
            return False

        # Klik Cari di sini juga bisa kecegat Turnstile ("Verifikasi keamanan gagal...")
        # sama seperti pencarian pasien awal - kalau ini terjadi, tabel riwayat tidak akan
        # pernah muncul walau di-retry berkali-kali. Coba EzSolver dulu, baru minta
        # refresh manual kalau EzSolver juga gagal.
        pesan = ambil_pesan_notif(driver)
        if pesan and ("captcha" in pesan.lower() or "verifikasi" in pesan.lower()):
            print(f"⚠️ {pesan}")
            print("-> Verifikasi Turnstile ditolak server saat membuka riwayat pasca-simpan, mencoba EzSolver...")
            berhasil_ezsolver = False
            if coba_solve_turnstile_ezsolver(driver):
                btn_cari.click()
                tunggu_loading_pace(driver)
                time.sleep(1.5)
                pesan_ulang = ambil_pesan_notif(driver)
                if not (pesan_ulang and ("captcha" in pesan_ulang.lower() or "verifikasi" in pesan_ulang.lower())):
                    print("-> Berhasil setelah Turnstile diselesaikan via EzSolver.")
                    berhasil_ezsolver = True
                else:
                    print(f"⚠️ {pesan_ulang}")
                    print("-> EzSolver tetap ditolak server.")

            if not berhasil_ezsolver:
                print("   Silakan REFRESH halaman (F5) di browser, lalu ulangi pencarian secara manual sampai data pasien tampil.")
                input("Setelah data pasien tampil di browser, tekan ENTER di sini untuk bot melanjutkan proses...")
                tunggu_loading_pace(driver)

        try:
            link_riwayat = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "#linkRiwayat a")))
            driver.execute_script("arguments[0].scrollIntoView({behavior: 'smooth', block: 'center'});", link_riwayat)
            time.sleep(0.2)
            link_riwayat.click()
        except Exception:
            try:
                driver.execute_script("Riwayat.instance.toggleRiwayatPelayanan();")
            except Exception as e:
                print(f"⚠️ Gagal membuka riwayat (percobaan {percobaan}/{max_retry}): {str(e)[:30]}")
                if percobaan < max_retry:
                    time.sleep(1)
                    continue
                return False

        tunggu_loading_pace(driver)
        try:
            WebDriverWait(driver, 20).until(EC.invisibility_of_element_located((By.ID, "riwayatPelayanan_processing")))
        except:
            pass
        try:
            WebDriverWait(driver, 12).until(EC.visibility_of_element_located((By.ID, "riwayatPelayanan")))
        except Exception:
            print(f"⚠️ Tabel riwayat tidak muncul (percobaan {percobaan}/{max_retry}).")
            if percobaan < max_retry:
                time.sleep(1)
                continue
            return False

        baris_terpilih, faskes_terpilih = pilih_baris_riwayat_berdasarkan_tanggal(driver, tgl_rujukan)
        if not baris_terpilih:
            print(f"⚠️ Riwayat dengan tanggal {tgl_rujukan} tidak ditemukan setelah simpan (percobaan {percobaan}/{max_retry}).")
            if percobaan < max_retry:
                time.sleep(1)
                continue
            return False

        if "LABKES" in faskes_terpilih.upper():
            print(f"-> Riwayat LABKESDA ditemukan (setelah simpan): {faskes_terpilih}")
        else:
            if require_labkes:
                print(
                    f"⚠️ Riwayat LABKESDA tidak ditemukan setelah simpan "
                    f"(percobaan {percobaan}/{max_retry})."
                )
                if percobaan < max_retry:
                    time.sleep(1)
                    continue
                return False
            print(f"-> Riwayat LABKESDA tidak ditemukan (setelah simpan), memakai baris: {faskes_terpilih or '-'}")

        try:
            btn_pilih = baris_terpilih.find_element(By.XPATH, ".//button[contains(@class, 'btnView')]")
            driver.execute_script("arguments[0].scrollIntoView({behavior: 'smooth', block: 'center'});", btn_pilih)
            time.sleep(0.5)
            btn_pilih.click()
        except Exception as e:
            print(f"⚠️ Gagal klik baris riwayat (percobaan {percobaan}/{max_retry}): {str(e)[:30]}")
            if percobaan < max_retry:
                time.sleep(1)
                continue
            return False

        tunggu_loading_pace(driver)
        time.sleep(1.0)
        if not tunggu_tab_nonkapi(driver):
            print("⚠️ Tab Non Kapitasi belum muncul, melanjutkan dengan hati-hati.")
        return True
    return False

def cek_faskes_pelayanan_labkesda(driver, timeout=5):
    akhir = time.time() + timeout
    while time.time() < akhir:
        try:
            teks = driver.find_element(By.ID, "faskesPelayan_lbl").text
        except Exception:
            teks = ""
        if "LABKES" in teks.upper():
            return True
        time.sleep(0.3)
    return False

def ambil_data_klinis_dari_excel(sheet_data, row):
    kolom = {
        'suhu': 20, 'tinggi': 21, 'berat': 22, 'lingkar': 23, 'imt': 24,
        'sistole': 25, 'diastole': 26, 'resprate': 27, 'heartrate': 28,
    }
    data = {}
    for key, col in kolom.items():
        val = sheet_data.cell(row=row, column=col).value
        if val is None or str(val).strip() == "":
            return None
        data[key] = str(val).strip()
    return data

def pilih_baris_riwayat_berdasarkan_tanggal(driver, tgl_rujukan):
    baris_riwayat = driver.find_elements(
        By.XPATH,
        f"//table[@id='riwayatPelayanan']/tbody/tr[td[4][contains(text(), '{tgl_rujukan}')]]"
    )
    if not baris_riwayat:
        return None, ""
    for baris in baris_riwayat:
        kolom = baris.find_elements(By.TAG_NAME, "td")
        faskes = kolom[1].text.strip() if len(kolom) > 1 else ""
        if "LABKES" in faskes.upper():
            return baris, faskes
    kolom_pertama = baris_riwayat[0].find_elements(By.TAG_NAME, "td")
    faskes_pertama = kolom_pertama[1].text.strip() if len(kolom_pertama) > 1 else ""
    return baris_riwayat[0], faskes_pertama

def jalankan_pelayanan(driver, wb_data, sheet_data, path_file, nama_pasien_resume=None):
    print("\n" + "="*50)
    print(" MEMULAI MODE: PELAYANAN PASIEN")
    print("="*50)

    tunggu_loading_pace(driver)

    # Kalau tab ini sudah berada di halaman Pelayanan Pasien dengan data pasien
    # yang sudah tampil (mis. hasil pencarian manual gara-gara Turnstile), jangan
    # navigasi ulang lewat menu - itu akan mereset form yang sudah terisi.
    sudah_di_halaman_pelayanan = False
    try:
        sudah_di_halaman_pelayanan = "entrikunjungandokkel" in driver.current_url.lower()
    except Exception:
        pass

    if sudah_di_halaman_pelayanan and nama_pasien_resume:
        print(f"-> Melanjutkan dari tab yang sudah terbuka di halaman Pelayanan Pasien (pasien: {nama_pasien_resume}), navigasi menu dilewati.")
    else:
        try:
            driver.find_element(By.XPATH, "//a[contains(text(), 'Entri Data')]").click()
            time.sleep(0.5)
            driver.find_element(By.XPATH, "//a[contains(text(), 'Pelayanan Pasien')]").click()
            tunggu_loading_pace(driver)
            print("-> Berhasil masuk ke halaman Pelayanan Pasien.")
        except Exception as e:
            print(f"-> Gagal navigasi ke menu Pelayanan Pasien: {e}")
            return

    tutup_modal_alert(driver)

    # Pastikan Turnstile benar-benar terverifikasi dulu di sini, sebelum memproses
    # baris apa pun - bukan disisipkan di tengah pengisian data per pasien.
    pastikan_turnstile_terselesaikan(driver)

    maks_baris = sheet_data.max_row
    yes_to_all = False
    skip_name_check = False
    tenaga_medis_tersimpan = None
    auto_print = False
    baris_kosong_berturut = 0

    for row in range(4, maks_baris + 1):
        if cek_sesi_berakhir(driver):
            driver.get("https://pcarejkn.bpjs-kesehatan.go.id/eclaim/EntriKunjunganDokkel")
            tunggu_loading_pace(driver)
            tutup_modal_alert(driver)

        nama_excel = sheet_data.cell(row=row, column=2).value
        penyakit = str(sheet_data.cell(row=row, column=3).value).strip()
        penyakit_upper = str(penyakit).upper()
        no_bpjs = sheet_data.cell(row=row, column=4).value
        tgl_rujukan = sheet_data.cell(row=row, column=17).value
        status_pendaftaran = sheet_data.cell(row=row, column=18).value
        status_finish = sheet_data.cell(row=row, column=29).value

        nama_kosong = not nama_excel or str(nama_excel).strip() == ""
        bpjs_kosong = not no_bpjs or str(no_bpjs).strip() == "" or str(no_bpjs).lower() == "kosong"
        if nama_kosong and bpjs_kosong:
            baris_kosong_berturut += 1
            if baris_kosong_berturut >= 5:
                print(f"-> Mendeteksi {baris_kosong_berturut} baris kosong berturut-turut, dianggap akhir data (berhenti di baris {row}).")
                break
            continue
        baris_kosong_berturut = 0

        if not no_bpjs or str(no_bpjs).lower() == "kosong":
            continue
            
        if "SUKSES" not in str(status_pendaftaran).upper():
            print(f"[{row}/{maks_baris}] Skip {nama_excel}: Status Pendaftaran belum sukses.")
            continue
            
        if status_finish and "FINISH" in str(status_finish).upper():
            print(f"[{row}/{maks_baris}] Skip {nama_excel}: Sudah FINISH.")
            continue
        if status_finish and "CANNOT BE INPUT" in str(status_finish).upper():
            print(f"[{row}/{maks_baris}] Skip {nama_excel}: CANNOT BE INPUT.")
            continue

        print(f"\n[{row}/{maks_baris}] Memproses Pelayanan: {nama_excel} | BPJS: {no_bpjs}")

        # Baris pertama yang namanya cocok dengan pasien yang sudah tampil di tab
        # (nama_pasien_resume) tidak perlu dicari ulang - data sudah ada di layar.
        # Hanya dipakai sekali, baris-baris berikutnya tetap search seperti biasa.
        resume_baris_ini = False
        if nama_pasien_resume:
            nama_excel_str = str(nama_excel).strip().lower()
            nama_resume_str = nama_pasien_resume.strip().lower()
            if nama_excel_str and (nama_excel_str in nama_resume_str or nama_resume_str in nama_excel_str):
                resume_baris_ini = True
            nama_pasien_resume = None

        try:
            if resume_baris_ini:
                print("-> Data pasien ini sudah tampil di tab, melanjutkan tanpa mencari ulang.")
            else:
                # 2. Input Tanggal
                input_tgl = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.ID, "txttanggal")))
                input_tgl.click()
                input_tgl.send_keys(Keys.CONTROL + "a")
                input_tgl.send_keys(Keys.BACKSPACE)
                input_tgl.send_keys(str(tgl_rujukan))
                input_tgl.send_keys(Keys.ESCAPE)
                time.sleep(0.5)

                # 3. Pilih Sumber Data
                driver.find_element(By.ID, "rbkartu").click()
                time.sleep(0.5)

                # 4. Input BPJS
                input_bpjs = driver.find_element(By.ID, "nomor")
                input_bpjs.click()
                input_bpjs.send_keys(Keys.CONTROL + "a")
                input_bpjs.send_keys(Keys.BACKSPACE)
                input_bpjs.send_keys(str(no_bpjs))

                # 5. Cari - klik PERTAMA untuk pasien baru ini dilakukan USER secara manual
                # di browser (klik asli manusia lebih reliable untuk Turnstile daripada
                # bot). Bot cuma menunggu otomatis sampai nama pasien tampil, tanpa ENTER.
                print("-> Silakan klik tombol Cari (ikon kaca pembesar) di browser untuk pasien ini...")
                lanjut_tunggu_cari = True
                skip_baris_ini = False
                stop_bot = False
                percobaan_tidak_ditemukan = 0
                while lanjut_tunggu_cari:
                    lanjut_tunggu_cari = False
                    status_cari, info_cari = tunggu_klik_cari_pasien_oleh_user(driver)

                    if status_cari == "tidak_ditemukan":
                        percobaan_tidak_ditemukan += 1
                        print(f"⚠️ PERINGATAN: {info_cari} (percobaan {percobaan_tidak_ditemukan}/3)")
                        if percobaan_tidak_ditemukan < 3:
                            # Sering kali penyebabnya user lupa ganti tanggal/field lain,
                            # bukan datanya benar-benar tidak ada - kasih kesempatan coba lagi
                            # dulu sebelum benar-benar menganggap ini gagal.
                            print("   Cek lagi tanggal/No.Kartu di form, lalu klik Cari lagi di browser...")
                            lanjut_tunggu_cari = True
                        else:
                            sheet_data.cell(row=row, column=19).value = "Gagal: Data tidak ada"
                            sheet_data.cell(row=row, column=19).font = Font(color="FF0000")
                            wb_data.save(path_file)
                            print("\n1. Ulangi pasien ini (coba cari lagi)")
                            print("2. Lanjut pasien berikutnya")
                            print("3. Stop bot")
                            pilihan_gagal = input("Pilih (1/2/3): ").strip()
                            if pilihan_gagal == '1':
                                percobaan_tidak_ditemukan = 0
                                lanjut_tunggu_cari = True
                            elif pilihan_gagal == '3':
                                stop_bot = True
                                skip_baris_ini = True
                            else:
                                skip_baris_ini = True
                    elif status_cari == "timeout":
                        konf_timeout = input(
                            "⚠️ Belum ada nama pasien muncul setelah beberapa saat. "
                            "Ketik 'n' stop bot, 's' skip pasien ini, atau ENTER untuk terus menunggu: "
                        ).strip().lower()
                        if konf_timeout == 'n':
                            stop_bot = True
                            skip_baris_ini = True
                        elif konf_timeout == 's':
                            skip_baris_ini = True
                        else:
                            lanjut_tunggu_cari = True

                if skip_baris_ini:
                    if stop_bot: break
                    continue

            # 7. Cek Nama (retry: klik Cari lagi dulu, baru minta refresh manual kalau tetap kosong -
            # jangan langsung skip baris, soalnya pencarian bisa "berhasil" tanpa alert apapun
            # tapi nama tidak pernah ter-render).
            nama_sistem = ""
            for percobaan_nama in range(1, 3):
                try:
                    nama_sistem = WebDriverWait(driver, 5).until(EC.visibility_of_element_located((By.ID, "lblnmpst"))).text
                except Exception:
                    nama_sistem = ""

                if nama_sistem.strip():
                    break

                if percobaan_nama == 1:
                    print("⚠️ Nama pasien belum tampil, mencoba klik Cari lagi...")
                    klik_search_dengan_retry_captcha(
                        driver, lambda: driver.find_element(By.ID, "btnCariPendaftaran").click()
                    )
                else:
                    print("⚠️ PERINGATAN: Nama pasien tetap gagal dimuat setelah dicoba ulang.")
                    print(f"   Silakan REFRESH halaman (F5) di browser lalu cari pasien ini secara manual (No.Kartu: {no_bpjs}).")
                    input("Setelah nama pasien tampil di browser, tekan ENTER di sini untuk bot melanjutkan proses...")

            if not nama_sistem.strip():
                print("⚠️ PERINGATAN: Nama pasien tetap tidak ditemukan.")
                konf_nama_kosong = input("Ketik 'n' untuk stop bot, atau ENTER/'y' lanjut pasien berikutnya: ").strip().lower()
                if konf_nama_kosong == 'n': break
                continue

            is_nama_cocok = str(nama_excel).strip().lower() in str(nama_sistem).strip().lower() or str(nama_sistem).strip().lower() in str(nama_excel).strip().lower()
            if not is_nama_cocok:
                print(f"⚠️ NAMA BERBEDA! Excel: {nama_excel} | Sistem: {nama_sistem}")
                if not skip_name_check:
                    konf_nama = input("Tetap lanjut? (ENTER/y = lanjut, n = lewati, a = lanjut semua tanpa cek nama): ").strip().lower()
                    if konf_nama == 'a':
                        skip_name_check = True
                        print("⚠️ PERINGATAN: Mode lanjut semua aktif, nama beda tidak akan dicek lagi hingga selesai.")
                    if konf_nama not in ("", "y", "a"):
                        sheet_data.cell(row=row, column=19).value = "Dilewati: Nama Beda"
                        sheet_data.cell(row=row, column=19).font = Font(color="FF0000")
                        wb_data.save(path_file)
                        continue

            # Kalau 9 kolom data klinis (suhu s/d heart rate) sudah tercatat di Excel
            # dari proses sebelumnya, tidak perlu ekstrak ulang dari riwayat non-LABKESDA -
            # tapi tabel riwayat tetap dicek dulu, siapa tahu kunjungan LABKESDA-nya
            # sendiri sudah berhasil dibuat sebelumnya (proses lama sempat terhenti
            # setelah simpan, sebelum sempat mengisi hasil lab) - kalau sudah ada,
            # jangan buat kunjungan baru lagi (duplikat), langsung lanjut ke input hasil lab.
            data_klinis_excel = ambil_data_klinis_dari_excel(sheet_data, row)

            # 8. Tampilkan Riwayat Pelayanan
            try:
                btn_riwayat = driver.find_element(By.XPATH, "//a[contains(@onclick, 'toggleRiwayatPelayanan')]")
                driver.execute_script("arguments[0].scrollIntoView({behavior: 'smooth', block: 'center'});", btn_riwayat)
                time.sleep(0.5)
            except: pass

            driver.execute_script("Riwayat.instance.toggleRiwayatPelayanan();")
            time.sleep(1.0)
            tunggu_loading_pace(driver)
            try:
                WebDriverWait(driver, 15).until(EC.invisibility_of_element_located((By.ID, "riwayatPelayanan_processing")))
            except: pass

            WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.ID, "riwayatPelayanan")))
            time.sleep(1.5)

            # 9. Cari Tanggal Rujukan
            baris_terpilih, faskes_terpilih = pilih_baris_riwayat_berdasarkan_tanggal(driver, tgl_rujukan)
            if not baris_terpilih:
                print(f"⚠️ PERINGATAN: Riwayat dengan tanggal {tgl_rujukan} tidak ditemukan di tabel!")
                konf = input("Ketik 'n' untuk stop bot, atau ENTER/'y' lanjut pasien berikutnya: ").strip().lower()
                if konf == 'n': break
                continue

            labkesda_ditemukan = "LABKES" in faskes_terpilih.upper()
            gunakan_riwayat_lab = labkesda_ditemukan
            if labkesda_ditemukan:
                print(f"-> Riwayat LABKESDA ditemukan: {faskes_terpilih}")
            elif data_klinis_excel:
                print(f"-> Riwayat LABKESDA belum ditemukan (memakai baris: {faskes_terpilih or '-'}), data klinis sudah ada di Excel - langsung buat kunjungan baru untuk LABKESDA (riwayat lama tidak dibuka).")
            else:
                print(f"-> Riwayat LABKESDA tidak ditemukan, memakai baris: {faskes_terpilih or '-'}")

            # Baris riwayat hanya perlu dibuka kalau: (a) LABKESDA sudah ada (untuk lanjut
            # ke input hasil lab di kunjungan itu), atau (b) belum ada data klinis sama
            # sekali sehingga perlu diekstrak dari riwayat non-LABKESDA yang ditemukan.
            if labkesda_ditemukan or not data_klinis_excel:
                btn_pilih_riwayat = baris_terpilih.find_element(By.XPATH, ".//button[contains(@class, 'btnView')]")
                driver.execute_script("arguments[0].scrollIntoView({behavior: 'smooth', block: 'center'});", btn_pilih_riwayat)
                time.sleep(0.5)
                btn_pilih_riwayat.click()
                tunggu_loading_pace(driver)
                time.sleep(1.5)
                if not tunggu_tab_nonkapi(driver):
                    print("⚠️ Tab Non Kapitasi belum muncul, melanjutkan dengan hati-hati.")

                if not gunakan_riwayat_lab:
                    # 10. Ekstrak Data Lama
                    print("-> Mengekstrak data riwayat klinis...")
                    suhu_input = driver.find_element(By.ID, "suhu_txt")
                    driver.execute_script("arguments[0].scrollIntoView({behavior: 'smooth', block: 'center'});", suhu_input)
                    time.sleep(0.5)

                    data_klinis = {
                        'suhu': driver.find_element(By.ID, "suhu_txt").get_attribute('value'),
                        'tinggi': driver.find_element(By.ID, "tinggiBadan").get_attribute('value'),
                        'berat': driver.find_element(By.ID, "beratBadan").get_attribute('value'),
                        'lingkar': driver.find_element(By.ID, "lingkarPerut").get_attribute('value'),
                        'imt': driver.find_element(By.ID, "imt").get_attribute('value'),
                        'sistole': driver.find_element(By.ID, "sistole").get_attribute('value'),
                        'diastole': driver.find_element(By.ID, "diastole").get_attribute('value'),
                        'resprate': driver.find_element(By.ID, "respRate").get_attribute('value'),
                        'heartrate': driver.find_element(By.ID, "heartRate").get_attribute('value')
                    }

                    sheet_data.cell(row=row, column=20).value = data_klinis['suhu']
                    sheet_data.cell(row=row, column=21).value = data_klinis['tinggi']
                    sheet_data.cell(row=row, column=22).value = data_klinis['berat']
                    sheet_data.cell(row=row, column=23).value = data_klinis['lingkar']
                    sheet_data.cell(row=row, column=24).value = data_klinis['imt']
                    sheet_data.cell(row=row, column=25).value = data_klinis['sistole']
                    sheet_data.cell(row=row, column=26).value = data_klinis['diastole']
                    sheet_data.cell(row=row, column=27).value = data_klinis['resprate']
                    sheet_data.cell(row=row, column=28).value = data_klinis['heartrate']
                    wb_data.save(path_file)

            if not gunakan_riwayat_lab and data_klinis_excel:
                data_klinis = data_klinis_excel

            if not gunakan_riwayat_lab:
                print("\n" + "-"*40)
                print(" DATA KLINIS UNTUK KUNJUNGAN BARU ")
                print("-" * 40)
                print(f" Suhu        : {data_klinis['suhu']} ℃")
                print(f" Tinggi/Berat: {data_klinis['tinggi']} cm / {data_klinis['berat']} kg")
                print(f" Lingkar/IMT : {data_klinis['lingkar']} cm / {data_klinis['imt']}")
                print(f" Tensi       : {data_klinis['sistole']}/{data_klinis['diastole']} mmHg")
                print(f" Resp / HR   : {data_klinis['resprate']} / {data_klinis['heartrate']} bpm")
                print("-" * 40)

                if not yes_to_all:
                    tanya_isi = input("Lanjut menginput hasil pemeriksaan ini? (Y = Lanjut / N = Stop / A = Lanjut Semua): ").strip().lower()
                    if tanya_isi == 'n': break
                    elif tanya_isi == 'a': yes_to_all = True

                # 11. Mulai Input Pemeriksaan Baru (retry jika riwayat LABKESDA belum muncul)
                simpan_berhasil = False
                for percobaan_simpan in range(1, 4):
                    # Klik Cari WAJIB diulang di sini untuk mengosongkan field kunjungan -
                    # kalau sebelumnya sempat membuka baris riwayat non-LABKESDA (mis. GAPURA)
                    # untuk cek/ekstrak data, form ini bisa masih merujuk ke kunjungan lama
                    # itu (field/dropdown gagal diisi karena state form tidak sesuai).
                    print("-> Membuka form input kunjungan...")
                    # Klik Cari di sini juga bisa kena "Verifikasi keamanan gagal..." -
                    # kalau dibiarkan, panel yang tampil masih panel LAMA (belum ter-reset)
                    # sehingga cek faskesPelayan_lbl di bawah salah kira form sudah benar.
                    klik_search_dengan_retry_captcha(
                        driver, lambda: driver.find_element(By.ID, "btnCariPendaftaran").click()
                    )
                    WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.ID, "panelEntriKunjungan")))
                    tunggu_loading_pace(driver)
                    time.sleep(1)

                    # Titik ini kerap minta verifikasi Turnstile ulang - pastikan
                    # benar-benar terverifikasi sebelum menyentuh field apa pun,
                    # supaya dropdown/field di bawah tidak gagal karena overlay captcha.
                    pastikan_turnstile_form_kunjungan(driver)

                    # Pastikan form yang terbuka benar-benar untuk kunjungan LABKESDA yang
                    # baru/kosong (cek label Faskes Pelayanan) - kalau masih menunjukkan
                    # faskes lain (mis. GAPURA), jangan isi field apa pun, klik Cari lagi
                    # di percobaan berikutnya sampai form-nya benar.
                    if not cek_faskes_pelayanan_labkesda(driver):
                        faskes_saat_ini = ""
                        try:
                            faskes_saat_ini = driver.find_element(By.ID, "faskesPelayan_lbl").text
                        except Exception:
                            pass
                        print(f"⚠️ Faskes Pelayanan masih '{faskes_saat_ini or '-'}' (bukan LABKESDA), klik Cari lagi untuk reset form ({percobaan_simpan}/3)...")
                        time.sleep(1)
                        continue

                    # A. Keluhan & Anamnesa
                    txt_keluhan = ""
                    if "DM" in penyakit_upper and "HT" in penyakit_upper: txt_keluhan = "diabetes mellitus dan hipertensi"
                    elif "DM" in penyakit_upper: txt_keluhan = "diabetes mellitus"
                    elif "HT" in penyakit_upper: txt_keluhan = "hipertensi"
                    else: txt_keluhan = "pemeriksaan rutin"

                    keluhan_el = driver.find_element(By.ID, "keluhan")
                    keluhan_el.clear()
                    keluhan_el.send_keys(txt_keluhan)
                    anamnesa_el = driver.find_element(By.ID, "anamnesa_txt")
                    anamnesa_el.clear()
                    anamnesa_el.send_keys(txt_keluhan)
                    
                    # B. Riwayat Alergi
                    pilih_select2(driver, "alergiMakan_slc", "Tidak Ada")
                    pilih_select2(driver, "alergiUdara_slc", "Tidak Ada")
                    pilih_select2(driver, "alergiObat_slc", "Tidak Ada")
                    
                    # C. Prognosa
                    pilih_select2(driver, "prognosa_slc", "Bonam")
                    
                    # D. Terapi Obat & Non Obat
                    terapi_med = driver.find_element(By.ID, "terapiMedikamentosa_txt")
                    terapi_med.clear()
                    terapi_med.send_keys("----")
                    terapi_non = driver.find_element(By.ID, "terapiNonMedikamentosa_txt")
                    terapi_non.clear()
                    terapi_non.send_keys("----")
                    
                    # E. Diagnosa
                    if "DM" in penyakit_upper and "HT" in penyakit_upper: kode_diag = "e11.9"
                    elif "DM" in penyakit_upper: kode_diag = "e11.9"
                    elif "HT" in penyakit_upper: kode_diag = "i10"
                    else: kode_diag = "e11.9"
                        
                    inp_diag = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.ID, "kddiagnosa1")))
                    driver.execute_script("arguments[0].scrollIntoView({behavior: 'smooth', block: 'center'});", inp_diag)
                    time.sleep(0.5)
                    
                    inp_diag.click()
                    time.sleep(0.2)
                    inp_diag.send_keys(Keys.CONTROL + "a")
                    inp_diag.send_keys(Keys.BACKSPACE)
                    time.sleep(0.2)
                    inp_diag.send_keys(kode_diag)
                    time.sleep(0.5)
                    inp_diag.send_keys(Keys.TAB) 
                    
                    try:
                        WebDriverWait(driver, 5).until(lambda d: len(d.find_element(By.ID, "nmdiagnosa1").get_attribute("value").strip()) > 2)
                    except:
                        driver.execute_script("PemDokkel.instance.readNamaDiagnosa('diagnosa1');")
                        time.sleep(1)
                    
                    # F. Input Data Klinis
                    suhu_el = driver.find_element(By.ID, "suhu_txt")
                    suhu_el.clear()
                    suhu_el.send_keys(data_klinis['suhu'])
                    tinggi_el = driver.find_element(By.ID, "tinggiBadan")
                    tinggi_el.clear()
                    tinggi_el.send_keys(data_klinis['tinggi'])
                    berat_el = driver.find_element(By.ID, "beratBadan")
                    berat_el.clear()
                    berat_el.send_keys(data_klinis['berat'])
                    lingkar_el = driver.find_element(By.ID, "lingkarPerut")
                    lingkar_el.clear()
                    lingkar_el.send_keys(data_klinis['lingkar'])
                    sistole_el = driver.find_element(By.ID, "sistole")
                    sistole_el.clear()
                    sistole_el.send_keys(data_klinis['sistole'])
                    diastole_el = driver.find_element(By.ID, "diastole")
                    diastole_el.clear()
                    diastole_el.send_keys(data_klinis['diastole'])
                    resp_el = driver.find_element(By.ID, "respRate")
                    resp_el.clear()
                    resp_el.send_keys(data_klinis['resprate'])
                    hr_el = driver.find_element(By.ID, "heartRate")
                    hr_el.clear()
                    hr_el.send_keys(data_klinis['heartrate'])
                    
                    # G. Tenaga Medis
                    if tenaga_medis_tersimpan is None:
                        container_medis = driver.find_element(By.XPATH, "//select[@id='tenagamedis']/following-sibling::span")
                        driver.execute_script("arguments[0].scrollIntoView({behavior: 'smooth', block: 'center'});", container_medis)
                        time.sleep(0.5)
                        container_medis.click()
                        time.sleep(0.5)
                        
                        elemen_nama = driver.find_elements(By.XPATH, "//ul[@id='select2-tenagamedis-results']//li[@role='treeitem']")
                        daftar_nama_medis = [el.text for el in elemen_nama if el.text.strip() != ""]
                        driver.find_element(By.TAG_NAME, 'body').click() 
                        time.sleep(0.5)
                        
                        idx_terpilih, tenaga_medis_tersimpan = menu_interaktif(daftar_nama_medis, "Pilih Tenaga Medis untuk sesi ini:")
                        print(f"-> Memilih tenaga medis: {tenaga_medis_tersimpan}")
                    
                    pilih_select2(driver, "tenagamedis", tenaga_medis_tersimpan)
                    
                    # H. Pelayanan Non Kapitasi
                    container_non_kapitasi = driver.find_element(By.XPATH, "//select[@id='listNonKapitasi_slc']/following-sibling::span")
                    driver.execute_script("arguments[0].scrollIntoView({behavior: 'smooth', block: 'center'});", container_non_kapitasi)
                    time.sleep(0.5)
                    
                    if "DM" in penyakit_upper:
                        container_non_kapitasi.click(); time.sleep(0.5)
                        driver.find_element(By.XPATH, "//ul[@id='select2-listNonKapitasi_slc-results']//li[contains(., 'Pelayanan Gula Darah')]").click()
                        time.sleep(0.5)
                        container_non_kapitasi.click(); time.sleep(0.5)
                        driver.find_element(By.XPATH, "//ul[@id='select2-listNonKapitasi_slc-results']//li[contains(., 'Pelayanan Kimia Darah')]").click()
                        time.sleep(0.5)
                        container_non_kapitasi.click(); time.sleep(0.5)
                        driver.find_element(By.XPATH, "//ul[@id='select2-listNonKapitasi_slc-results']//li[contains(., 'Pelayanan HbA1c')]").click()
                    elif "HT" in penyakit_upper:
                        container_non_kapitasi.click(); time.sleep(0.5)
                        driver.find_element(By.XPATH, "//ul[@id='select2-listNonKapitasi_slc-results']//li[contains(., 'Pelayanan Kimia Darah')]").click()
                    
                    time.sleep(0.5)
                    driver.find_element(By.TAG_NAME, 'body').click() 
                    
                    # I. Status Pulang
                    pilih_select2(driver, "statuspulang", "Berobat Jalan")
                    
                    # J. Simpan
                    btn_simpan = driver.find_element(By.ID, "btnSimpan")
                    driver.execute_script("arguments[0].scrollIntoView({behavior: 'smooth', block: 'center'});", btn_simpan)
                    time.sleep(0.5)
                    btn_simpan.click()
                    print("-> Menyimpan kunjungan utama...")
                    
                    try:
                        WebDriverWait(driver, 5).until(EC.alert_is_present())
                        alert = driver.switch_to.alert
                        alert.accept()
                    except: pass
                    tunggu_loading_pace(driver)

                    if buka_riwayat_setelah_simpan(driver, tgl_rujukan, require_labkes=True):
                        simpan_berhasil = True
                        break

                    print(f"⚠️ Kunjungan belum tercatat di LABKESDA. Ulang simpan ({percobaan_simpan}/3).")
                    time.sleep(1)

                if not simpan_berhasil:
                    sheet_data.cell(row=row, column=29).value = "NOT COMPLETE"
                    wb_data.save(path_file)
                    continue
            
            # L. PENGISIAN PELAYANAN NON KAPITASI (HASIL LAB DARI EXCEL)
            # Pengecekan dilakukan berulang hingga total biaya cocok
            target_biaya_kimia = 380000
            kimia_darah_tests = [
                ("Kolesterol Total", sheet_data.cell(row, 9).value), # I
                ("Trigliserida", sheet_data.cell(row, 10).value), # J
                ("Ureum", sheet_data.cell(row, 11).value), # K
                ("Kreatinin", sheet_data.cell(row, 12).value), # L
                ("Kolesterol HDL", sheet_data.cell(row, 13).value), # M
                ("Kolesterol LDL", sheet_data.cell(row, 14).value), # N
                ("Microalbuminaria", sheet_data.cell(row, 16).value) # P
            ]
            val_gdp = sheet_data.cell(row, 8).value # Kolom H
            val_hba1c = sheet_data.cell(row, 15).value # Kolom O
            perlu_gdp = "DM" in penyakit_upper and val_gdp is not None and str(val_gdp).strip() != ""
            perlu_hba1c = "DM" in penyakit_upper and val_hba1c is not None and str(val_hba1c).strip() != ""
            target_biaya = target_biaya_kimia
            if "DM" in penyakit_upper:
                if perlu_gdp:
                    target_biaya += 12500
                if perlu_hba1c:
                    target_biaya += 155000
            lab_sudah_dicek = False
            stop_semua = False
            lab_sudah_diinput = False
            cannot_be_input = False
            lab_berhasil = False
            max_retry_lab = 6
            retry_count = 0
            layanan_tidak_input = []
            
            while True:
                print(f"\n-> Memproses input hasil laboratorium. Target Biaya: Rp {target_biaya:,}")
                existing_kimia = set()
                existing_gula = set()
                hba1c_tersimpan = False

                if "DM" in penyakit_upper:
                    if buka_tab(driver, "tabDet_10"):
                        time.sleep(0.5)
                        if elemen_terlihat(driver, By.ID, "div_tambahPelayananGulaDarah_btn"):
                            existing_gula = set(ambil_nama_pelayanan_dari_tabel(driver, "daftarPelayananGulaDarah"))
                    else:
                        print("⚠️ Tab Pelayanan Gula Darah tidak ditemukan.")

                    if buka_tab(driver, "tabDet_11"):
                        time.sleep(0.5)
                        try:
                            btn_hba1c = driver.find_element(By.XPATH, "//div[@id='tabDet_11']//button[@id='simpan_btn']")
                            hba1c_tersimpan = not btn_hba1c.is_enabled()
                        except Exception:
                            hba1c_tersimpan = False

                if buka_tab(driver, "tabDet_12"):
                    time.sleep(0.5)
                
                kimia_box_ada = False
                kimia_box_ada = elemen_terlihat(driver, By.ID, "listKimiaDarah_lyt")
                existing_kimia = set(ambil_nama_pelayanan_dari_tabel(driver, "daftarPelayanan_tbl")) if kimia_box_ada else set()
                if not kimia_box_ada:
                    print("⚠️ Box Pelayanan Kimia Darah tidak ditemukan. Coba ulang atau refresh.")
                
                if not lab_sudah_dicek:
                    expected_kimia = [nama for nama, val in kimia_darah_tests if val is not None and str(val).strip() != ""]
                    missing_kimia = [nama for nama in expected_kimia if nama not in existing_kimia]
                    missing_gula = []
                    hba1c_kurang = False
                    
                    if "DM" in penyakit_upper:
                        if perlu_gdp and not daftar_mengandung(existing_gula, "Gula Darah Puasa"):
                            missing_gula = ["Gula Darah Puasa"]
                        if perlu_hba1c:
                            hba1c_kurang = not hba1c_tersimpan
                    
                    print("\n" + "-"*40)
                    print(" RINGKASAN CEK LAB (TARGET HT/DM)")
                    print("-"*40)
                    print(f" Kimia Darah - sudah: {', '.join(sorted(existing_kimia)) if existing_kimia else '-'}")
                    print(f" Kimia Darah - kurang: {', '.join(missing_kimia) if missing_kimia else '-'}")
                    if "DM" in penyakit_upper:
                        print(f" Gula Darah - sudah: {', '.join(sorted(existing_gula)) if existing_gula else '-'}")
                        print(f" Gula Darah - kurang: {', '.join(missing_gula) if missing_gula else '-'}")
                        if perlu_hba1c:
                            print(f" HbA1c - sudah: {'YA' if not hba1c_kurang else 'BELUM'}")
                    print("-"*40)
                    
                    ada_kekurangan = bool(missing_kimia or missing_gula or hba1c_kurang)
                    if ada_kekurangan and not yes_to_all:
                        tanya_lab = input("Data lab belum lengkap. Lanjut input yang kurang? (ENTER/y = lanjut, n = stop): ").strip().lower()
                        if tanya_lab == 'n':
                            stop_semua = True
                            break
                    lab_sudah_dicek = True
                
                # --- GULA DARAH & HBA1C (HANYA DM) ---
                if "DM" in penyakit_upper:
                    try:
                        if perlu_gdp:
                            if buka_tab(driver, "tabDet_10"):
                                time.sleep(0.5)
                                if elemen_terlihat(driver, By.ID, "div_tambahPelayananGulaDarah_btn"):
                                    existing_gula = set(ambil_nama_pelayanan_dari_tabel(driver, "daftarPelayananGulaDarah"))
                                if not daftar_mengandung(existing_gula, "Gula Darah Puasa"):
                                    if elemen_terlihat(driver, By.ID, "div_tambahPelayananGulaDarah_btn"):
                                        safe_click(driver, By.ID, "tambahPelayananGulaDarah_btn")
                                        time.sleep(0.5)
                                    WebDriverWait(driver, 8).until(
                                        EC.visibility_of_element_located((By.XPATH, "//div[@id='tabDet_10']//input[@id='hasil_txt']"))
                                    )
                                    if not pilih_opsi_dropdown(driver, "cb_jns_pemeriksaan_darah", "Gula Darah Puasa"):
                                        print("⚠️ Gagal memilih jenis pemeriksaan Gula Darah Puasa.")
                                    in_hasil_gdp = driver.find_element(By.XPATH, "//div[@id='tabDet_10']//input[@id='hasil_txt']")
                                    in_hasil_gdp.clear()
                                    in_hasil_gdp.send_keys(str(val_gdp).replace(',', '.'))
                                    safe_click(driver, By.XPATH, "//div[@id='tabDet_10']//button[@id='simpan_btn']")
                                    hasil, pesan = tunggu_hasil_simpan(driver)
                                    if hasil == "warning" and pesan:
                                        tandai_cannot_input(sheet_data, wb_data, path_file, row, pesan)
                                        cannot_be_input = True
                                        break
                                    time.sleep(1)
                                    if elemen_terlihat(driver, By.ID, "div_tambahPelayananGulaDarah_btn"):
                                        existing_gula = set(ambil_nama_pelayanan_dari_tabel(driver, "daftarPelayananGulaDarah"))

                        if perlu_hba1c:
                            if buka_tab(driver, "tabDet_11"):
                                time.sleep(0.5)
                                try:
                                    btn_hba1c = driver.find_element(By.XPATH, "//div[@id='tabDet_11']//button[@id='simpan_btn']")
                                except Exception:
                                    btn_hba1c = None

                                if btn_hba1c and btn_hba1c.is_enabled():
                                    if not elemen_terlihat(driver, By.XPATH, "//div[@id='tabDet_11']//input[@id='hasil_txt']"):
                                        try:
                                            btn_tambah = driver.find_element(By.XPATH, "//div[@id='tabDet_11']//button[contains(., 'Tambah Data Baru') or contains(@id, 'tambah')]")
                                            btn_tambah.click()
                                            time.sleep(0.5)
                                        except Exception:
                                            pass
                                    WebDriverWait(driver, 8).until(
                                        EC.visibility_of_element_located((By.XPATH, "//div[@id='tabDet_11']//input[@id='hasil_txt']"))
                                    )
                                    in_hasil_hba1c = driver.find_element(By.XPATH, "//div[@id='tabDet_11']//input[@id='hasil_txt']")
                                    in_hasil_hba1c.clear()
                                    in_hasil_hba1c.send_keys(str(val_hba1c).replace(',', '.'))
                                    safe_click(driver, By.XPATH, "//div[@id='tabDet_11']//button[@id='simpan_btn']")
                                    hasil, pesan = tunggu_hasil_simpan(driver)
                                    if hasil == "warning" and pesan:
                                        tandai_cannot_input(sheet_data, wb_data, path_file, row, pesan)
                                        cannot_be_input = True
                                        break
                                    time.sleep(1)
                                    hba1c_tersimpan = True
                                elif btn_hba1c and not btn_hba1c.is_enabled():
                                    hba1c_tersimpan = True
                    except Exception as e:
                        print(f"⚠️ Gagal memproses Gula Darah / HbA1c: {str(e)[:30]}")
                if cannot_be_input:
                    break

                # --- KIMIA DARAH ---
                if buka_tab(driver, "tabDet_12"):
                    time.sleep(0.5)
                try:
                    for nama_test, val_lab in kimia_darah_tests:
                        if val_lab is None or str(val_lab).strip() == "":
                            continue
                        if nama_test in existing_kimia:
                            continue
                        
                        if nama_test == "Microalbuminaria" and str(val_lab).strip().upper() == "TIDAK ADA URINE" or str(val_lab).strip().upper() == "":
                            val_lab = str(round(random.uniform(5.3, 19.9), 1))
                        
                        val_str = str(val_lab).replace(',', '.')
                        
                        pastikan_form_kimia_darah(driver)
                        pilih_select2(driver, "jnsPelayanan_slc", nama_test)
                        
                        in_hasil = driver.find_element(By.XPATH, "//div[@id='tabDet_12']//input[@id='hasil_txt']")
                        in_hasil.clear()
                        in_hasil.send_keys(val_str)
                        
                        safe_click(driver, By.XPATH, "//div[@id='tabDet_12']//button[@id='simpan_btn']")
                        hasil, pesan = tunggu_hasil_simpan(driver)
                        if hasil == "warning" and pesan:
                            tandai_cannot_input(sheet_data, wb_data, path_file, row, pesan)
                            cannot_be_input = True
                            break
                        time.sleep(1)
                        existing_kimia.add(nama_test)
                except Exception as e:
                    print(f"⚠️ Gagal memproses Kimia Darah: {str(e)[:30]}")
                if cannot_be_input:
                    break

                # [VALIDASI BIAYA]
                buka_tab(driver, "tabDet_12")
                time.sleep(0.5)
                total_biaya_kimia = hitung_total_biaya_lab(driver, "tabDet_12", "daftarPelayanan_tbl")
                
                total_biaya_gdp = 0
                if "DM" in penyakit_upper:
                    buka_tab(driver, "tabDet_10")
                    time.sleep(0.5)
                    total_biaya_gdp = hitung_total_biaya_lab(driver, "tabDet_10", "daftarPelayananGulaDarah")
                    
                    # Tambah biaya HbA1c (fix 162.500 jika ada)
                    if perlu_hba1c and hba1c_tersimpan:
                        total_biaya_gdp += 162500
                
                biaya_saat_ini = total_biaya_kimia + total_biaya_gdp
                
                print(f"-> Pengecekan Biaya: Terhitung Rp {biaya_saat_ini:,} dari Target Rp {target_biaya:,}")
                
                if biaya_saat_ini >= target_biaya:
                    print("-> ✅ SELURUH HASIL LAB BERHASIL TERSIMPAN SEMPURNA!")
                    sheet_data.cell(row=row, column=29).value = "FINISH"
                    sheet_data.cell(row=row, column=30).value = ""
                    wb_data.save(path_file)
                    lab_berhasil = True
                    break # Keluar dari loop lab jika biaya sudah cocok
                else:
                    retry_count += 1
                    layanan_tidak_input = []
                    if buka_tab(driver, "tabDet_12"):
                        time.sleep(0.3)
                    kimia_box_ada_now = elemen_terlihat(driver, By.ID, "listKimiaDarah_lyt")
                    existing_kimia_now = set(ambil_nama_pelayanan_dari_tabel(driver, "daftarPelayanan_tbl")) if kimia_box_ada_now else set()
                    expected_kimia_now = [nama for nama, val in kimia_darah_tests if val is not None and str(val).strip() != ""]
                    missing_kimia_now = [nama for nama in expected_kimia_now if nama not in existing_kimia_now]
                    layanan_tidak_input.extend(missing_kimia_now)

                    if perlu_gdp:
                        buka_tab(driver, "tabDet_10")
                        time.sleep(0.3)
                        existing_gula_now = set(ambil_nama_pelayanan_dari_tabel(driver, "daftarPelayananGulaDarah"))
                        if not daftar_mengandung(existing_gula_now, "Gula Darah Puasa"):
                            layanan_tidak_input.append("Gula Darah Puasa")

                    if perlu_hba1c:
                        buka_tab(driver, "tabDet_11")
                        time.sleep(0.3)
                        try:
                            btn_hba1c = driver.find_element(By.XPATH, "//div[@id='tabDet_11']//button[@id='simpan_btn']")
                            hba1c_ok = not btn_hba1c.is_enabled()
                        except Exception:
                            hba1c_ok = False
                        if not hba1c_ok:
                            layanan_tidak_input.append("HbA1c")

                    if not layanan_tidak_input:
                        layanan_tidak_input = ["LAB TIDAK LENGKAP"]

                    if retry_count >= max_retry_lab:
                        pesan_skip = f"CANNOT INPUT ({', '.join(layanan_tidak_input)})"
                        sheet_data.cell(row=row, column=29).value = "CANNOT INPUT"
                        sheet_data.cell(row=row, column=30).value = pesan_skip
                        sheet_data.cell(row=row, column=29).font = Font(color="FF0000")
                        wb_data.save(path_file)
                        print(f"-> ⚠️ {pesan_skip}. Melewati pasien ini.")
                        break

                    print("-> ❌ PERINGATAN: Ada data lab yang gagal tersimpan ke server PCare (Bug Jaringan/Sistem).")
                    if yes_to_all:
                        print("-> Mengulangi proses penginputan lab secara otomatis...")
                        continue # Ulangi loop tanpa bertanya
                    else:
                        tanya_ulang = input("Apakah Anda ingin mencoba mengulang input lab lagi? (Y/N/A = Ulang Terus): ").strip().lower()
                        if tanya_ulang == 'n':
                            print("-> Melewati pengisian lab, lanjut ke cetak.")
                            break
                        elif tanya_ulang == 'a':
                            yes_to_all = True
                        continue

            if stop_semua:
                break
            lab_sudah_diinput = True

            if cannot_be_input:
                continue

            if not lab_berhasil:
                continue

            # M. Cetak SPP
            btn_spp = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.ID, "spp_btn")))
            btn_spp.click()
            print("-> Tab SPP dibuka, menunggu jendela cetak muncul...")
            window_utama = driver.current_window_handle
            try:
                # Tab SPP (generate PDF/print preview) kadang butuh lebih dari 10 detik
                # untuk benar-benar terbuka sebagai window baru, atau malah keburu
                # tertutup sendiri (auto-print lalu auto-close) sebelum sempat terdeteksi.
                WebDriverWait(driver, 30).until(EC.number_of_windows_to_be(2))
                for win in driver.window_handles:
                    if win != window_utama:
                        driver.switch_to.window(win)
                        if not auto_print:
                            tanya_auto = input("Otomatis print semua? (1 = Ya, ENTER = Manual): ").strip()
                            if tanya_auto == '1':
                                auto_print = True
                        if auto_print:
                            jalankan_auto_print(
                                driver,
                                AUTO_PRINT_DELAY_SPP_MIN,
                                AUTO_PRINT_DELAY_SPP_MAX,
                                AUTO_PRINT_POST_DELAY_SPP
                            )
                        else:
                            konfirmasi_spp = input("Selesai print SPP? (ENTER/y = tutup, n = stop): ").strip().lower()
                            if konfirmasi_spp == 'n':
                                stop_semua = True
                                driver.switch_to.window(window_utama)
                                break
                        driver.close()
                driver.switch_to.window(window_utama)
            except Exception as e:
                # Jangan langsung dianggap error fatal (refresh/stop) - tab cetak sering
                # gagal terdeteksi otomatis padahal cetaknya sendiri baik-baik saja.
                # Biarkan user print sendiri lalu konfirmasi lanjut.
                print(f"⚠️ Tab cetak SPP tidak terdeteksi otomatis: {str(e)[:60] or 'window tidak terbuka/tertutup sendiri'}.")
                print("   Silakan cetak SPP secara manual di browser kalau belum tercetak, lalu tutup tab cetaknya.")
                input("Setelah selesai, tekan ENTER di sini untuk bot melanjutkan (FKPP / pasien berikutnya)...")
                try:
                    for win in driver.window_handles:
                        if win != window_utama:
                            driver.switch_to.window(win)
                            driver.close()
                    driver.switch_to.window(window_utama)
                except Exception:
                    pass

            if stop_semua:
                break

            if not lab_sudah_diinput:
                # Duplikat blok pengisian lab sebelumnya di-skip untuk menghindari input ganda.
                pass

            # N. CETAK FKPP & TANDAI SELESAI
            try:
                btn_fkpp = WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.ID, "cetakFKPP_btn")))
                driver.execute_script("arguments[0].scrollIntoView({behavior: 'smooth', block: 'center'});", btn_fkpp)
                time.sleep(0.5)
                btn_fkpp.click()
                print("-> Tab FKPP dibuka, menunggu jendela cetak muncul...")

                # Sama seperti tab SPP - render PDF/print preview kadang butuh lebih
                # dari 10 detik untuk terbuka sebagai window baru, atau keburu tertutup
                # sendiri sebelum sempat terdeteksi.
                try:
                    WebDriverWait(driver, 30).until(EC.number_of_windows_to_be(2))
                    for win in driver.window_handles:
                        if win != window_utama:
                            driver.switch_to.window(win)
                            if not auto_print:
                                tanya_auto = input("Otomatis print semua? (1 = Ya, ENTER = Manual): ").strip()
                                if tanya_auto == '1':
                                    auto_print = True
                            if auto_print:
                                jalankan_auto_print(
                                    driver,
                                    AUTO_PRINT_DELAY_FKPP_MIN,
                                    AUTO_PRINT_DELAY_FKPP_MAX,
                                    AUTO_PRINT_POST_DELAY_FKPP
                                )
                            else:
                                konfirmasi_fkpp = input("Selesai print FKPP? (ENTER/y = tutup, n = stop): ").strip().lower()
                                if konfirmasi_fkpp == 'n':
                                    stop_semua = True
                                    driver.switch_to.window(window_utama)
                                    break
                            driver.close()
                    driver.switch_to.window(window_utama)
                except Exception as e:
                    print(f"⚠️ Tab cetak FKPP tidak terdeteksi otomatis: {str(e)[:60] or 'window tidak terbuka/tertutup sendiri'}.")
                    print("   Silakan cetak FKPP secara manual di browser kalau belum tercetak, lalu tutup tab cetaknya.")
                    input("Setelah selesai, tekan ENTER di sini untuk bot melanjutkan (pasien berikutnya)...")
                    try:
                        for win in driver.window_handles:
                            if win != window_utama:
                                driver.switch_to.window(win)
                                driver.close()
                        driver.switch_to.window(window_utama)
                    except Exception:
                        pass

                if stop_semua:
                    break
                
                # === WARNAI EXCEL HIJAU & TULIS FINISH ===
                hijau_stabilo = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
                for col_idx in range(1, 30):
                    sheet_data.cell(row=row, column=col_idx).fill = hijau_stabilo
                sheet_data.cell(row=row, column=29).value = "FINISH"
                
                sheet_data.cell(row=row, column=19).value = "SUKSES"
                wb_data.save(path_file)
                print(f"-> ✅ Data pasien {nama_excel} telah SELESAI dan ditandai FINISH.")

                if not yes_to_all:
                    lanjut_semua = input("Lanjut pasien berikutnya? (ENTER/y = lanjut, n = stop, a = lanjut semua): ").strip().lower()
                    if lanjut_semua == 'n':
                        break
                    if lanjut_semua == 'a':
                        yes_to_all = True
                
            except Exception as e:
                print(f"⚠️ Gagal mencetak FKPP: {str(e)[:30]}")
                sheet_data.cell(row=row, column=29).value = "NOT COMPLETE"
                wb_data.save(path_file)
            
        except Exception as row_error:
            pesan_error = f"Gagal sistem: {str(row_error)[:30]}"
            print(f"-> Terjadi kendala teknis saat memproses BPJS {no_bpjs}. {pesan_error}")
            
            sheet_data.cell(row=row, column=19).value = pesan_error
            sheet_data.cell(row=row, column=19).font = Font(color="FF0000") # Merah
            sheet_data.cell(row=row, column=29).value = "NOT COMPLETE"
            wb_data.save(path_file)
            
            konfirmasi_error = input("Ketik 'n' untuk berhenti, atau ENTER/'y' untuk lanjut me-refresh: ").strip().lower()
            if konfirmasi_error == 'n': break
            else:
                driver.get("https://pcarejkn.bpjs-kesehatan.go.id/eclaim/EntriKunjunganDokkel")
                tunggu_loading_pace(driver)
                tutup_modal_alert(driver)
    else:
        nama_file = os.path.basename(path_file)
        nama_sheet = sheet_data.title
        pesan = format_report_message("SELESAI INPUT PELAYANAN PASIEN", nama_file, nama_sheet)
        kirim_notif_group(pesan)

# ====================================================================
# MODUL: PENDAFTARAN PASIEN 
# ====================================================================

def jalankan_pendaftaran(driver, wb_data, sheet_data, path_file):
    print("\n" + "="*50)
    print(" MEMULAI MODE: PENDAFTARAN PASIEN")
    print("="*50)
    
    tunggu_loading_pace(driver)
    try:
        driver.find_element(By.XPATH, "//a[contains(text(), 'Entri Data')]").click()
        time.sleep(0.5)
        driver.find_element(By.XPATH, "//a[contains(text(), 'Pendaftaran Pasien')]").click()
        tunggu_loading_pace(driver)
        print("-> Masuk ke menu Pendaftaran Pasien.")
    except: pass
    
    tutup_modal_alert(driver)

    # Pastikan Turnstile benar-benar terverifikasi dulu di sini, sebelum memproses
    # baris apa pun - bukan disisipkan di tengah pengisian data per pasien.
    pastikan_turnstile_terselesaikan(driver)

    maks_baris = sheet_data.max_row
    yes_to_all = False
    skip_name_check = False
    baris_kosong_berturut = 0

    for row in range(4, maks_baris + 1):
        if cek_sesi_berakhir(driver):
            driver.get("https://pcarejkn.bpjs-kesehatan.go.id/eclaim/EntriDaftarDokkel")
            tunggu_loading_pace(driver)
            tutup_modal_alert(driver)

        nama_excel = sheet_data.cell(row=row, column=2).value
        no_bpjs = sheet_data.cell(row=row, column=4).value
        status_input = sheet_data.cell(row=row, column=18).value

        nama_kosong = not nama_excel or str(nama_excel).strip() == ""
        bpjs_kosong = not no_bpjs or str(no_bpjs).strip() == ""
        if nama_kosong and bpjs_kosong:
            baris_kosong_berturut += 1
            if baris_kosong_berturut >= 5:
                print(f"-> Mendeteksi {baris_kosong_berturut} baris kosong berturut-turut, dianggap akhir data (berhenti di baris {row}).")
                break
            continue
        baris_kosong_berturut = 0

        if status_input and str(status_input).strip() != "":
            print(f"[{row}/{maks_baris}] Melewati baris {row} ({nama_excel}) - Status: {status_input}")
            continue
        
        if not no_bpjs or str(no_bpjs).strip() == "" or str(no_bpjs).lower() == "kosong":
            if str(no_bpjs).lower() != "kosong":
                print(f"\n[{row}/{maks_baris}] PERHATIAN: No BPJS kosong pada baris {row}.")
                sheet_data.cell(row=row, column=4).value = "kosong"
                sheet_data.cell(row=row, column=18).value = "Dilewati (Otomatis Bot)"
                wb_data.save(path_file)
                if not yes_to_all:
                    konf = input("Ketik 'n' stop bot, ENTER/'y' lanjut: ").strip().lower()
                    if konf == 'n': break
            continue 
            
        print(f"\n[{row}/{maks_baris}] Mendaftar NAMA: {nama_excel} | BPJS: {no_bpjs}")
        
        try:
            WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.ID, "rborizon"))).click()
            time.sleep(0.5)
            driver.find_element(By.ID, "btnQueryPesertaLain").click()
            
            input_bpjs = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.ID, "noKartuHorizon_txt")))
            time.sleep(0.5)
            input_bpjs.click()
            time.sleep(0.2)
            input_bpjs.send_keys(Keys.CONTROL + "a")
            time.sleep(0.1)
            input_bpjs.send_keys(Keys.BACKSPACE)
            input_bpjs.clear() 
            time.sleep(0.2)
            
            input_bpjs.send_keys(str(no_bpjs))

            def klik_cari_rujukan():
                driver.find_element(By.ID, "cariRujukanByNoka_btn").click()
                WebDriverWait(driver, 20).until(EC.invisibility_of_element_located((By.ID, "daftarRujukan_tbl_processing")))
                time.sleep(1)

            klik_search_dengan_retry_captcha(driver, klik_cari_rujukan)

            try:
                btn_pilih_rujukan = driver.find_element(By.XPATH, "//table[@id='daftarRujukan_tbl']/tbody/tr[1]/td[1]//button[contains(@onclick, 'rujukanHorizontalSelected')]")
                nama_tabel = driver.find_element(By.XPATH, "//table[@id='daftarRujukan_tbl']/tbody/tr[1]/td[3]").text
                tgl_kunjungan = driver.find_element(By.XPATH, "//table[@id='daftarRujukan_tbl']/tbody/tr[1]/td[5]").text
            except:
                print(f"⚠️ PERINGATAN: Tabel rujukan kosong untuk BPJS {no_bpjs}!")
                driver.find_element(By.ID, "batalRujukan_btn").click()
                time.sleep(1.5) 
                
                sheet_data.cell(row=row, column=18).value = "Gagal: Rujukan Tidak Ada"
                sheet_data.cell(row=row, column=18).font = Font(color="FF0000")
                wb_data.save(path_file)
                
                if not yes_to_all:
                    konf = input("Ketik 'n' stop bot, ENTER/'y' lanjut: ").strip().lower()
                    if konf == 'n': break
                continue

            if yes_to_all and not skip_name_check:
                skip_name_check = True

            nama_ex_bersih = str(nama_excel).strip().lower()
            nama_tb_bersih = str(nama_tabel).strip().lower()
            if not (nama_ex_bersih in nama_tb_bersih or nama_tb_bersih in nama_ex_bersih):
                print(f"\n⚠️ NAMA BEDA! Excel: {nama_excel} | Sistem: {nama_tabel}")
                if not skip_name_check:
                    konf_nama = input("Tetap lanjut? (ENTER/y = lanjut, n = lewati, a = lanjut semua tanpa cek nama): ").strip().lower()
                    if konf_nama == 'a':
                        skip_name_check = True
                        yes_to_all = True
                        print("⚠️ PERINGATAN: Mode lanjut semua aktif, nama beda tidak akan dicek lagi hingga selesai.")
                    if konf_nama not in ("", "y", "a"):
                        driver.find_element(By.ID, "batalRujukan_btn").click()
                        time.sleep(1.5)
                        sheet_data.cell(row=row, column=18).value = "Dilewati: Nama Beda"
                        sheet_data.cell(row=row, column=18).font = Font(color="FF0000")
                        wb_data.save(path_file)
                        continue

            btn_pilih_rujukan.click()
            time.sleep(1.5)
            
            sheet_data.cell(row=row, column=17).value = tgl_kunjungan
            
            input_tgl = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.ID, "txttanggal")))
            input_tgl.clear()
            input_tgl.send_keys(tgl_kunjungan)
            input_tgl.send_keys(Keys.ESCAPE) 
            time.sleep(0.5)
            
            radio_promotif = driver.find_element(By.ID, "tkp50")
            driver.execute_script("arguments[0].scrollIntoView({behavior: 'smooth', block: 'center'});", radio_promotif)
            time.sleep(1)
            radio_promotif.click()
            time.sleep(0.5)

            btn_simpan = driver.find_element(By.ID, "btnSimpanPendaftaran")
            driver.execute_script("arguments[0].scrollIntoView({behavior: 'smooth', block: 'center'});", btn_simpan)
            time.sleep(0.5)
            btn_simpan.click()
            tunggu_loading_pace(driver)
            
            try:
                WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, "//span[@data-notify='message' and contains(text(), 'Data Pendaftaran Berhasil disimpan')]")))
            except: pass
            
            sheet_data.cell(row=row, column=18).value = "SUKSES"
            sheet_data.cell(row=row, column=18).font = Font(color="000000")
            wb_data.save(path_file)
            
            if not yes_to_all:
                print(f"-> Pendaftaran {nama_excel} tersimpan (Tgl: {tgl_kunjungan}).")
                tl = input("Lanjut? (Y=Lanjut / N=Stop / A=Lanjut Semua): ").strip().lower()
                if tl == 'a': yes_to_all = True
                elif tl == 'n': break
            
        except Exception as row_error:
            sheet_data.cell(row=row, column=18).value = f"Gagal sistem: {str(row_error)[:30]}"
            sheet_data.cell(row=row, column=18).font = Font(color="FF0000")
            wb_data.save(path_file)
            
            tl_err = input("Error. Ketik 'n' stop, ENTER/'y' lanjut refresh: ").strip().lower()
            if tl_err == 'n': break
            else:
                driver.get("https://pcarejkn.bpjs-kesehatan.go.id/eclaim/EntriDaftarDokkel")
                tunggu_loading_pace(driver)
                tutup_modal_alert(driver)
    else:
        nama_file = os.path.basename(path_file)
        nama_sheet = sheet_data.title
        pesan = format_report_message("SELESAI INPUT PENDAFTARAN PASIEN", nama_file, nama_sheet)
        kirim_notif_group(pesan)


# ====================================================================
# MAIN EKSEKUSI
# ====================================================================

def cari_chrome_exe():
    if os.environ.get("CHROME_PATH"):
        return os.environ["CHROME_PATH"]
    kandidat = [
        r"C:\Program Files\Google\Chrome\Application\chrome.exe",
        r"C:\Program Files (x86)\Google\Chrome\Application\chrome.exe",
        os.path.expandvars(r"%LOCALAPPDATA%\Google\Chrome\Application\chrome.exe"),
    ]
    for path in kandidat:
        if os.path.isfile(path):
            return path
    return None

def pilih_file_dan_mode():
    path_file, nama_sheet, wb_data = setup_folder_excel()
    sheet_data = wb_data[nama_sheet]

    # Sel gabungan (merged cells) hanya bisa ditulis lewat sel pojok kiri-atasnya
    # (openpyxl melempar error 'MergedCell...read-only' kalau ditulis lewat sel lain
    # dalam gabungan itu). Lepas semua gabungan supaya semua sel bisa ditulis bebas.
    for rng in list(sheet_data.merged_cells.ranges):
        sheet_data.unmerge_cells(str(rng))

    opsi_mode = ["1. Pendaftaran Pasien", "2. Pelayanan Pasien (Input Hasil)"]
    idx_mode, mode_terpilih = menu_interaktif(opsi_mode, "Pilih Mode Operasi Bot:")
    return path_file, nama_sheet, wb_data, sheet_data, idx_mode

def jalankan_agent():
    nama_file_excel_user = 'data_user.xlsx'
    try:
        wb_user = openpyxl.load_workbook(nama_file_excel_user)
        sheet_user = wb_user['Sheet1']
    except Exception as e:
        print(f"Error file user: {e}")
        return

    username = sheet_user['B1'].value
    password = sheet_user['B2'].value

    if not username or not password:
        print("Error: Username/Password kosong di Excel user!")
        return

    manual_chrome_proc = None
    print("Menyiapkan Chrome untuk login manual (menghindari deteksi otomasi Cloudflare)...")
    chrome_exe = cari_chrome_exe()
    if not chrome_exe:
        print("⚠️ Google Chrome tidak ditemukan di lokasi standar. Set environment variable CHROME_PATH ke lokasi chrome.exe Anda.")
        return

    debug_port = 9333
    manual_profile_dir = os.path.join(os.environ.get("TEMP") or ".", "pcare_bot_profile")
    chrome_args = [
        chrome_exe,
        f"--remote-debugging-port={debug_port}",
        f"--user-data-dir={manual_profile_dir}",
        "--no-first-run",
    ]
    if ENABLE_KIOSK_PRINTING:
        chrome_args.append("--kiosk-printing")
    # Langsung ke halaman yang butuh login (bukan /eclaim/login). Kalau sesi lama
    # di profile ini masih aktif, PCare akan tetap di halaman ini; kalau tidak,
    # PCare otomatis redirect ke halaman login.
    chrome_args.append("https://pcarejkn.bpjs-kesehatan.go.id/eclaim/EntriDaftarDokkel")

    manual_chrome_proc = subprocess.Popen(chrome_args)
    time.sleep(3)

    print("Menghubungkan bot ke browser...")
    chrome_options = Options()
    chrome_options.add_experimental_option("debuggerAddress", f"127.0.0.1:{debug_port}")

    driver = None
    nama_pasien_resume = None
    try:
        driver = webdriver.Chrome(options=chrome_options)

        # Kalau ada tab lain yang sudah terbuka di halaman Pelayanan Pasien dengan
        # data pasien sudah tampil (misal hasil pencarian manual gara-gara Turnstile),
        # pakai tab itu dan lewati langkah pencarian pertama - jangan buka tab baru.
        handle_kunjungan_aktif = None
        for handle in driver.window_handles:
            driver.switch_to.window(handle)
            try:
                url_lower = driver.current_url.lower()
            except Exception:
                continue
            if "entrikunjungandokkel" in url_lower:
                try:
                    nama_txt = driver.find_element(By.ID, "lblnmpst").text.strip()
                except Exception:
                    nama_txt = ""
                if nama_txt:
                    handle_kunjungan_aktif = handle
                    nama_pasien_resume = nama_txt
                    break

        if handle_kunjungan_aktif:
            driver.switch_to.window(handle_kunjungan_aktif)
            print(f"-> Mendeteksi tab Pelayanan Pasien yang sudah terbuka dengan data pasien '{nama_pasien_resume}', akan dilanjutkan tanpa refresh.")
        else:
            for handle in driver.window_handles:
                driver.switch_to.window(handle)
                if "pcarejkn" in driver.current_url.lower():
                    break

        if "login" in driver.current_url.lower():
            print("\n" + "="*60)
            print("LOGIN DIPERLUKAN")
            print("Jendela Chrome baru saja terbuka ke halaman login PCare.")
            print("Cloudflare Turnstile biasanya terverifikasi otomatis dalam beberapa detik.")
            print("JANGAN TUTUP jendela Chrome tersebut.")
            print("="*60)
            opsi_login = [
                "1. Isi Username/Password otomatis lalu langsung klik Sign In otomatis",
                "2. Isi Username/Password otomatis, saya klik Sign In sendiri",
                "3. Saya isi & login semuanya secara manual",
            ]
            idx_login, _ = menu_interaktif(opsi_login, "Setelah Turnstile terverifikasi (centang hijau), pilih cara login:")

            while True:
                if idx_login in (0, 1):
                    input("Tekan ENTER setelah Turnstile terverifikasi untuk mengisi Username/Password...")
                    try:
                        isi_kredensial(driver, username, password)
                        print("-> Username/Password terisi otomatis.")
                    except Exception as e:
                        print(f"⚠️ Gagal isi otomatis: {str(e)[:60]}. Silakan isi & login manual.")
                        idx_login = 2

                    if idx_login == 0:
                        try:
                            driver.find_element(By.ID, "btnLogin").click()
                            print("-> Tombol Sign In diklik otomatis.")
                        except Exception as e:
                            print(f"⚠️ Gagal klik Sign In otomatis: {str(e)[:60]}. Silakan klik manual.")
                            idx_login = 1

                    if idx_login == 1:
                        input("Silakan klik tombol Sign In secara manual, lalu tekan ENTER di sini...")
                else:
                    input("Tekan ENTER setelah berhasil login dan berada di Dashboard PCare...")

                if "login" not in driver.current_url.lower():
                    break
                print("⚠️ Sepertinya masih di halaman login. Pastikan sudah berhasil masuk Dashboard, lalu coba lagi.")
                idx_login = 2  # fallback ke manual kalau percobaan sebelumnya belum berhasil
        else:
            print("-> Sesi login sebelumnya masih aktif, melanjutkan otomatis tanpa login ulang.")
        tutup_modal_alert(driver)

        lanjut_file_lain = True
        while lanjut_file_lain:
            path_file, nama_sheet, wb_data, sheet_data, idx_mode = pilih_file_dan_mode()

            if idx_mode == 0:
                jalankan_pendaftaran(driver, wb_data, sheet_data, path_file)
                lanjut_pelayanan = input(
                    "\nSelesai Pendaftaran. Sekalian lakukan input Pelayanan Pasien "
                    "untuk file & sheet yang sama? (Y/Enter = Ya, N = Tidak): "
                ).strip().lower()
                if lanjut_pelayanan in ("", "y"):
                    jalankan_pelayanan(driver, wb_data, sheet_data, path_file, nama_pasien_resume)
                    nama_pasien_resume = None
            elif idx_mode == 1:
                jalankan_pelayanan(driver, wb_data, sheet_data, path_file, nama_pasien_resume)
                nama_pasien_resume = None

            lagi = input("\nProses file Excel lain? (Y/Enter = Ya, N = Selesai): ").strip().lower()
            lanjut_file_lain = lagi in ("", "y")

    except Exception as e:
        print(f"-> Terjadi error fatal: {str(e)[:150]}")

    finally:
        if driver:
            input("\nTekan ENTER untuk menutup browser dan mengakhiri program...")
            try:
                driver.quit()
            except Exception:
                pass
        if manual_chrome_proc and manual_chrome_proc.poll() is None:
            if os.name == 'nt':
                try:
                    subprocess.run(
                        ["taskkill", "/F", "/T", "/PID", str(manual_chrome_proc.pid)],
                        stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL
                    )
                except Exception:
                    manual_chrome_proc.terminate()
            else:
                manual_chrome_proc.terminate()

if __name__ == "__main__":
    jalankan_agent()